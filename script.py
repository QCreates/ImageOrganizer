import openpyxl

FILE_NAME = "input"
FILE_TYPE = ".xlsx"
FILE_PATH = FILE_NAME + FILE_TYPE
SHEET_NAME = "Sheet1"

# Header names based on your sheet
IMAGES_HEADER = "Images"      # the first image column
FIRST_EXTRA_HEADER = "2"      # next image column
LAST_EXTRA_HEADER = "20"      # last image column


def is_blank_image_cell(ws, row, images_col):
    value = ws.cell(row=row, column=images_col).value
    return value in (None, "", " ")


# -----------------------------------------
# Load workbook
# -----------------------------------------
wb = openpyxl.load_workbook(FILE_PATH)
ws = wb[SHEET_NAME]

# Build header map
headers = {}
for col in range(1, ws.max_column + 1):
    v = ws.cell(row=1, column=col).value
    if v not in (None, "", " "):
        headers[str(v).strip()] = col

IMAGES_COL = headers[IMAGES_HEADER]

# Identify extra image columns
EXTRA_COLS = []
for name, col in headers.items():
    if name.isdigit():
        EXTRA_COLS.append(col)
print(EXTRA_COLS)

# Add new Image Index column at far right
INDEX_COL = ws.max_column + 1
ws.cell(row=1, column=INDEX_COL).value = "Image Index"

# -----------------------------------------
# Main loop
# -----------------------------------------
row = 2

while row <= ws.max_row:
    main_img = ws.cell(row=row, column=IMAGES_COL).value

    # Only process rows that have an image in IMAGES column
    if not main_img:
        row += 1
        continue

    # Collect all extra images from columns 2..20
    extra_images = []
    for col in EXTRA_COLS:
        v = ws.cell(row=row, column=col).value
        if v not in (None, "", " "):
            extra_images.append(v)
        

    # Nothing to do?
    if not extra_images:
        row += 1
        continue

    insert_row = row + 1

    img_index = 1
    # Place each extra image into an available or newly inserted row
    handle = ""
    for img in extra_images:
        # If we've passed the bottom → always insert a new row
        if insert_row > ws.max_row:
            ws.insert_rows(insert_row)
            ws.cell(row=insert_row, column=IMAGES_COL-2).value = handle

        # If next row's Images cell is NOT empty → new product starts → insert row
        elif not is_blank_image_cell(ws, insert_row, IMAGES_COL):
            ws.insert_rows(insert_row)
            ws.cell(row=insert_row, column=IMAGES_COL-2).value = handle
            #Insert handle
        else:
            handle = ws.cell(row=insert_row, column=IMAGES_COL-2).value
            print(handle)
        # Write image into the Images column ONLY
        ws.cell(row=insert_row, column=IMAGES_COL).value = img
        ws.cell(row=insert_row-1, column=INDEX_COL).value = img_index
        
        img_index += 1
        insert_row += 1
    # Add last index number
    ws.cell(row=insert_row-1, column=INDEX_COL).value = img_index
    print(f"Row ({row}/{ws.max_row})")
    row += 1


# -----------------------------------------
# Save file
# -----------------------------------------
wb.save(FILE_NAME + "_updated.xlsx")
print("Updated.")
